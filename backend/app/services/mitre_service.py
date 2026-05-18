import io
from app.extensions import db
from app.models.tactic import Tactic
from app.models.ttp import TTP

# ── default source URLs ────────────────────────────────────────────────────────

_DEFAULTS = {
    "enterprise": {
        "tactics":    "https://github.com/CyberCX-STA/PurpleOps-Deps/raw/master/attack.mitre/15.1/enterprise-attack-v15.1-tactics.xlsx",
        "techniques": "https://github.com/CyberCX-STA/PurpleOps-Deps/raw/master/attack.mitre/15.1/enterprise-attack-v15.1-techniques.xlsx",
    },
    "ics":    {"tactics": "", "techniques": ""},
    "mobile": {"tactics": "", "techniques": ""},
}

_SETTING_KEYS = {
    "enterprise": {"tactics": "mitre_tactics_url",        "techniques": "mitre_techniques_url"},
    "ics":        {"tactics": "mitre_ics_tactics_url",    "techniques": "mitre_ics_techniques_url"},
    "mobile":     {"tactics": "mitre_mobile_tactics_url", "techniques": "mitre_mobile_techniques_url"},
}


def _get_url(framework: str, component: str) -> str:
    from app.models.app_setting import AppSetting
    key     = _SETTING_KEYS[framework][component]
    default = _DEFAULTS[framework][component]
    stored  = AppSetting.get(key)
    return stored if stored else default


# ── xlsx download ──────────────────────────────────────────────────────────────

def _download_workbook(framework: str, component: str):
    import requests
    from openpyxl import load_workbook

    url = _get_url(framework, component)
    if not url:
        raise ValueError(f"No URL configured for {framework} {component}. Set it in Settings → TTP Library.")

    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    ws = wb.active
    rows = list(ws.rows)
    headers = [cell.value for cell in rows[0]]
    return rows[1:], headers


def _get(row, headers, col_name, default=""):
    try:
        idx = headers.index(col_name)
        val = row[idx].value
        return val if val is not None else default
    except (ValueError, IndexError):
        return default


# ── refresh tactics ────────────────────────────────────────────────────────────

def refresh_tactics(framework: str = "enterprise") -> int:
    rows, headers = _download_workbook(framework, "tactics")
    count = 0
    for row in rows:
        mitre_id = _get(row, headers, "ID")
        name     = _get(row, headers, "name")
        if not mitre_id or not name:
            continue
        tactic = Tactic.query.filter_by(mitre_id=mitre_id).first()
        if tactic:
            tactic.name      = name
            tactic.framework = framework
        else:
            tactic = Tactic(mitre_id=mitre_id, name=name, framework=framework)
            db.session.add(tactic)
        count += 1
    db.session.commit()
    return count


# ── refresh techniques ─────────────────────────────────────────────────────────

def refresh_techniques(framework: str = "enterprise") -> int:
    rows, headers = _download_workbook(framework, "techniques")

    tactic_map = {t.name.strip().lower(): t for t in Tactic.query.filter_by(framework=framework).all()}

    count = 0
    for row in rows:
        mitre_id = _get(row, headers, "ID")
        name     = _get(row, headers, "name")
        if not mitre_id or not name:
            continue

        description  = _get(row, headers, "description")
        tactics_str  = _get(row, headers, "tactics")
        platform_val = _get(row, headers, "platforms") or _get(row, headers, "platform")

        tactic_names = [t.strip() for t in tactics_str.split(",") if t.strip()]
        primary_tactic = tactic_names[0] if tactic_names else "Unknown"
        tactic_objs  = [tactic_map[n.lower()] for n in tactic_names if n.lower() in tactic_map]

        ttp = TTP.query.filter_by(mitre_id=mitre_id).first()
        if ttp:
            ttp.name        = name
            ttp.tactic      = primary_tactic
            ttp.description = description
            ttp.platform    = platform_val
            ttp.framework   = framework
        else:
            ttp = TTP(
                mitre_id=mitre_id,
                name=name,
                tactic=primary_tactic,
                description=description,
                platform=platform_val,
                framework=framework,
            )
            db.session.add(ttp)
            db.session.flush()

        ttp.tactics = tactic_objs
        count += 1

    db.session.commit()
    return count


# ── refresh all (or one) framework ────────────────────────────────────────────

def refresh_all(framework: str = None) -> dict:
    frameworks = [framework] if framework else ["enterprise", "ics", "mobile"]
    result = {}
    errors = {}
    for fw in frameworks:
        try:
            t = refresh_tactics(fw)
            te = refresh_techniques(fw)
            result[fw] = {"tactics_updated": t, "techniques_updated": te}
        except ValueError as e:
            errors[fw] = str(e)
    if errors:
        result["errors"] = errors
    return result
